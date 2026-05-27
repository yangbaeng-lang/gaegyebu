# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2026년 소득세 계산"

# ── 색상 정의 ──────────────────────────────────────────────
C_TITLE   = "1F3864"   # 진남색
C_BH      = "2E75B6"   # 병훈 헤더 (파랑)
C_AR      = "C55A11"   # 아름 헤더 (주황)
C_LABEL   = "D6E4F0"   # 구분 행 배경
C_LABEL_AR= "FCE4D6"
C_RESULT  = "FFF2CC"   # 결과 강조
C_WHITE   = "FFFFFF"
C_GRAY    = "F2F2F2"
C_RED     = "FF0000"

def ft(bold=False, size=10, color="000000", name="맑은 고딕"):
    return Font(bold=bold, size=size, color=color, name=name)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="595959")
    return Border(left=s, right=s, top=s, bottom=s)

def money(n):
    return f"{int(n):,}"

def pct(n):
    return f"{n:.1f}%"

def set_cell(ws, row, col, value, bold=False, size=10,
             fg=None, font_color="000000", align="center",
             border=True, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = ft(bold=bold, size=size, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fg:
        cell.fill = fill(fg)
    if border:
        cell.border = border_thin()
    if num_format:
        cell.number_format = num_format
    return cell

# ── 열 너비 ───────────────────────────────────────────────
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16

row = 1

# ═══════════════════════════════════════════════════════════
# 제목
# ═══════════════════════════════════════════════════════════
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="2026년 소득세 계산서 (연간 추정)")
c.font = Font(bold=True, size=14, color="FFFFFF", name="맑은 고딕")
c.fill = fill(C_TITLE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 32
row += 1

ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1,
    value="※ PS·PI는 연중 1회 기수령 금액 적용 | 월별 수당은 5개월 실적 × 2.4배 연환산 | 기준일: 2026-05-21")
c.font = Font(italic=True, size=9, color="595959", name="맑은 고딕")
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = fill("E9EFF7")
ws.row_dimensions[row].height = 18
row += 2

# ═══════════════════════════════════════════════════════════
# 1. 연간 총급여 추정
# ═══════════════════════════════════════════════════════════
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="① 연간 총급여 추정")
c.font = Font(bold=True, size=11, color="FFFFFF", name="맑은 고딕")
c.fill = fill(C_TITLE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[row].height = 22
row += 1

# 헤더
headers = ["구 분", "병훈 (원)", "아름 (원)", "비 고"]
header_fills = [C_LABEL, C_BH, C_AR, C_LABEL]
header_colors = ["000000", "FFFFFF", "FFFFFF", "000000"]
ws.merge_cells(f"A{row}:A{row}")  # 구분 (A)
for i, (h, f_, fc) in enumerate(zip(headers, header_fills, header_colors), 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, size=10, color=fc, name="맑은 고딕")
    c.fill = fill(f_)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws.row_dimensions[row].height = 20
row += 1

salary_items = [
    ("PS (성과급, 기수령)",      72_277_410,  68_192_370,  "연중 1회 지급"),
    ("PI",                      6_531_720,   6_162_560,   "연중 1회 지급"),
    ("기본급 (연환산)",           56_020_320,  50_495_424,  "×2.4"),
    ("고정시간외+교대+야간 수당", 13_248_264,  12_631_008,  "×2.4"),
    ("평일연장+휴일근무 수당",    8_022_408,   5_627_712,   "×2.4"),
    ("명절격려금 (2회)",          960_000,     0,           "설·추석"),
    ("기타 수당 (연환산)",        5_107_896,   4_013_753,   "×2.4"),
]

for i, (label, bh, ar, note) in enumerate(salary_items):
    bg = C_WHITE if i % 2 == 0 else C_GRAY
    set_cell(ws, row, 1, label,    fg=bg, align="left")
    set_cell(ws, row, 2, bh,       fg=bg, align="right", num_format="#,##0")
    set_cell(ws, row, 3, ar,       fg=bg, align="right", num_format="#,##0")
    set_cell(ws, row, 4, note,     fg=bg, align="center")
    row += 1

# 합계행
bh_total = 162_168_018
ar_total  = 147_122_827
set_cell(ws, row, 1, "연간 총급여 합계", bold=True, fg=C_RESULT, align="left")
set_cell(ws, row, 2, bh_total, bold=True, fg=C_RESULT, align="right", num_format="#,##0")
set_cell(ws, row, 3, ar_total, bold=True, fg=C_RESULT, align="right", num_format="#,##0")
set_cell(ws, row, 4, "", fg=C_RESULT)
ws.row_dimensions[row].height = 22
row += 2

# ═══════════════════════════════════════════════════════════
# 2. 소득세 계산 과정
# ═══════════════════════════════════════════════════════════
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="② 소득세 계산 과정")
c.font = Font(bold=True, size=11, color="FFFFFF", name="맑은 고딕")
c.fill = fill(C_TITLE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[row].height = 22
row += 1

# 헤더
for i, (h, f_, fc) in enumerate(zip(headers, header_fills, header_colors), 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, size=10, color=fc, name="맑은 고딕")
    c.fill = fill(f_)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws.row_dimensions[row].height = 20
row += 1

bh_근소공제 = 15_993_360
ar_근소공제 = 15_692_457
bh_근소금액 = bh_total - bh_근소공제   # 146,174,658
ar_근소금액 = ar_total - ar_근소공제   # 131,430,370
bh_인적공제 = 1_500_000
ar_인적공제 = 9_500_000
bh_과표 = bh_근소금액 - bh_인적공제    # 144,674,658
ar_과표 = ar_근소금액 - ar_인적공제    # 121,930,370
bh_산출 = 35_196_130
ar_산출 = 27_235_630
bh_세액공제 = 500_000
ar_세액공제 = 500_000
bh_결정 = bh_산출 - bh_세액공제       # 34,696,130
ar_결정 = ar_산출 - ar_세액공제       # 26,735,630
bh_지방 = round(bh_결정 * 0.1)
ar_지방 = round(ar_결정 * 0.1)

calc_items = [
    ("총급여",              bh_total,      ar_total,      ""),
    ("(-) 근로소득공제",    -bh_근소공제,   -ar_근소공제,   "1억 초과: 1,475만+초과×2%"),
    ("= 근로소득금액",      bh_근소금액,    ar_근소금액,    ""),
    ("(-) 인적공제",        -bh_인적공제,   -ar_인적공제,   "병훈:본인1명 | 아름:5명+장애인"),
    ("= 과세표준",          bh_과표,        ar_과표,        "35% 구간 해당 (8,800만~1.5억)"),
    ("산출세액 (×35%-1,544만)", bh_산출,    ar_산출,        ""),
    ("(-) 근로소득세액공제", -bh_세액공제,  -ar_세액공제,   "총급여 7,000만 초과 → 최대 50만"),
    ("= 결정세액 (소득세)",  bh_결정,       ar_결정,        ""),
    ("지방소득세 (×10%)",   bh_지방,        ar_지방,        ""),
]

important_rows = {"= 결정세액 (소득세)", "과세표준", "= 근로소득금액"}
for i, (label, bh, ar, note) in enumerate(calc_items):
    bg = C_RESULT if label.startswith("=") else (C_WHITE if i % 2 == 0 else C_GRAY)
    bold = label.startswith("=")
    set_cell(ws, row, 1, label, bold=bold, fg=bg, align="left")
    set_cell(ws, row, 2, bh,   bold=bold, fg=bg, align="right", num_format="#,##0")
    set_cell(ws, row, 3, ar,   bold=bold, fg=bg, align="right", num_format="#,##0")
    set_cell(ws, row, 4, note, bold=False, fg=bg, align="center")
    row += 1

row += 1

# ═══════════════════════════════════════════════════════════
# 3. 공제 항목 합계 (4대보험 포함)
# ═══════════════════════════════════════════════════════════
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="③ 전체 공제 항목 요약")
c.font = Font(bold=True, size=11, color="FFFFFF", name="맑은 고딕")
c.fill = fill(C_TITLE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[row].height = 22
row += 1

for i, (h, f_, fc) in enumerate(zip(headers, header_fills, header_colors), 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, size=10, color=fc, name="맑은 고딕")
    c.fill = fill(f_)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws.row_dimensions[row].height = 20
row += 1

bh_np   = 3_331_800
ar_np   = 3_331_800
bh_hi   = 5_748_857
ar_hi   = 5_215_505
bh_lt   = 744_477
ar_lt   = 675_408
bh_emp  = 1_459_512
ar_emp  = 1_324_106
bh_total_ded = bh_결정 + bh_지방 + bh_np + bh_hi + bh_lt + bh_emp
ar_total_ded = ar_결정 + ar_지방 + ar_np + ar_hi + ar_lt + ar_emp
bh_net  = bh_total - bh_total_ded
ar_net  = ar_total - ar_total_ded

deduct_items = [
    ("소득세",            bh_결정,       ar_결정,       "결정세액"),
    ("지방소득세",         bh_지방,       ar_지방,       "소득세×10%"),
    ("국민연금",          bh_np,         ar_np,         "상한: 617만×12×4.5%"),
    ("건강보험",          bh_hi,         ar_hi,         "총급여×3.545%"),
    ("장기요양보험",       bh_lt,         ar_lt,         "건보×12.95%"),
    ("고용보험",          bh_emp,        ar_emp,        "총급여×0.9%"),
]

for i, (label, bh, ar, note) in enumerate(deduct_items):
    bg = C_WHITE if i % 2 == 0 else C_GRAY
    set_cell(ws, row, 1, label, fg=bg, align="left")
    set_cell(ws, row, 2, bh,   fg=bg, align="right", num_format="#,##0")
    set_cell(ws, row, 3, ar,   fg=bg, align="right", num_format="#,##0")
    set_cell(ws, row, 4, note, fg=bg, align="center")
    row += 1

# 총공제 합계
set_cell(ws, row, 1, "총 공제액 합계", bold=True, fg="FF6B6B", font_color="FFFFFF", align="left")
set_cell(ws, row, 2, bh_total_ded, bold=True, fg="FF6B6B", font_color="FFFFFF", align="right", num_format="#,##0")
set_cell(ws, row, 3, ar_total_ded, bold=True, fg="FF6B6B", font_color="FFFFFF", align="right", num_format="#,##0")
set_cell(ws, row, 4, "", fg="FF6B6B")
ws.row_dimensions[row].height = 22
row += 1

# 실수령액
set_cell(ws, row, 1, "연간 실수령액", bold=True, fg="375623", font_color="FFFFFF", align="left")
set_cell(ws, row, 2, bh_net, bold=True, fg="375623", font_color="FFFFFF", align="right", num_format="#,##0")
set_cell(ws, row, 3, ar_net, bold=True, fg="375623", font_color="FFFFFF", align="right", num_format="#,##0")
set_cell(ws, row, 4, "", fg="375623")
ws.row_dimensions[row].height = 22
row += 2

# ═══════════════════════════════════════════════════════════
# 4. 실효세율 요약
# ═══════════════════════════════════════════════════════════
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="④ 실효세율 요약")
c.font = Font(bold=True, size=11, color="FFFFFF", name="맑은 고딕")
c.fill = fill(C_TITLE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[row].height = 22
row += 1

rate_headers = ["지 표", "병 훈", "아 름", "비 고"]
for i, (h, f_, fc) in enumerate(zip(rate_headers, header_fills, header_colors), 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, size=10, color=fc, name="맑은 고딕")
    c.fill = fill(f_)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()
ws.row_dimensions[row].height = 20
row += 1

bh_rate1 = bh_결정 / bh_total * 100
ar_rate1 = ar_결정 / ar_total * 100
bh_rate2 = (bh_결정 + bh_지방) / bh_total * 100
ar_rate2 = (ar_결정 + ar_지방) / ar_total * 100
bh_rate3 = bh_total_ded / bh_total * 100
ar_rate3 = ar_total_ded / ar_total * 100

rate_items = [
    ("적용 세율 구간",   "35%",    "35%",   "과표 8,800만~1.5억"),
    ("실효 소득세율",    f"{bh_rate1:.1f}%", f"{ar_rate1:.1f}%", "소득세/총급여"),
    ("실효세율 (지방 포함)", f"{bh_rate2:.1f}%", f"{ar_rate2:.1f}%", "(소득세+지방세)/총급여"),
    ("총 공제율",        f"{bh_rate3:.1f}%", f"{ar_rate3:.1f}%", "전체 공제/총급여"),
]

for i, (label, bh, ar, note) in enumerate(rate_items):
    bg = C_WHITE if i % 2 == 0 else C_GRAY
    bold = label in ["실효세율 (지방 포함)", "총 공제율"]
    set_cell(ws, row, 1, label, bold=bold, fg=bg, align="left")
    set_cell(ws, row, 2, bh,   bold=bold, fg=bg, align="center")
    set_cell(ws, row, 3, ar,   bold=bold, fg=bg, align="center")
    set_cell(ws, row, 4, note, fg=bg, align="center")
    row += 1

row += 1

# ═══════════════════════════════════════════════════════════
# 5. 주의사항
# ═══════════════════════════════════════════════════════════
ws.merge_cells(f"A{row}:E{row}")
c = ws.cell(row=row, column=1, value="⑤ 주의사항 및 추가 공제 가능 항목")
c.font = Font(bold=True, size=11, color="FFFFFF", name="맑은 고딕")
c.fill = fill("7B7B7B")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[row].height = 22
row += 1

notes = [
    "본 계산서는 추정치이며 실제 연말정산 결과와 다를 수 있습니다.",
    "연금저축·IRP 납입분 세액공제 미반영 → 납입 시 추가 절세 가능 (최대 148.5만원)",
    "의료비·교육비·기부금 세액공제 미반영 → 실제 지출에 따라 환급 가능",
    "주택청약저축·장기주택저당차입금 이자 공제 미반영",
    "연금보조·의료비지원·복리후생 포인트 중 일부는 비과세 항목일 수 있음",
    "건강보험료는 연말 보수월액 정산에 따라 실납부액이 달라질 수 있음",
    "2026년 사대보험 요율은 2025년 기준 적용 (확정 후 재계산 권장)",
]
for i, note in enumerate(notes):
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=f"  • {note}")
    c.font = Font(size=9, color="404040", name="맑은 고딕")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = fill(C_WHITE if i % 2 == 0 else C_GRAY)
    c.border = border_thin()
    ws.row_dimensions[row].height = 18
    row += 1

# ── 저장 ──────────────────────────────────────────────────
out_path = r"f:\One Drive\OneDrive - SK Hynix Inc\AI 작업폴더\가계부 홈페이지 제작\2026년_소득세_계산서.xlsx"
wb.save(out_path)
print(f"저장 완료: {out_path}")
